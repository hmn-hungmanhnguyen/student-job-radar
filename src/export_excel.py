from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment


ROOT_DIR = Path(__file__).resolve().parent.parent

CSV_FILE = ROOT_DIR / "data" / "jobs.csv"
EXCEL_FILE = ROOT_DIR / "data" / "jobs.xlsx"


def main():
    if not CSV_FILE.exists():
        print("No jobs.csv found. Run parse_jobs.py first.")
        return

    df = pd.read_csv(CSV_FILE)

    if "days_until_deadline" in df.columns:
        df["sort_deadline"] = pd.to_numeric(df["days_until_deadline"], errors="coerce")
        df = df.sort_values(
            by=["sort_deadline"],
            ascending=True,
            na_position="last"
        )
        df = df.drop(columns=["sort_deadline"])

    df.to_excel(EXCEL_FILE, index=False)

    workbook = load_workbook(EXCEL_FILE)
    sheet = workbook.active
    sheet.title = "Public Jobs"

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    widths = {
        "A": 25,   # source_name
        "B": 70,   # title
        "C": 16,   # job_type
        "D": 18,   # deadline
        "E": 20,   # days_until_deadline
        "F": 22,   # opportunity_status
        "G": 35,   # email
        "H": 35,   # published
        "I": 70,   # source_url
        "J": 70,   # pdf_url
        "K": 45,   # public_note
        "L": 100,  # summary
    }

    for col, width in widths.items():
        sheet.column_dimensions[col].width = width

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    workbook.save(EXCEL_FILE)

    print(f"Saved public Excel file to: {EXCEL_FILE}")


if __name__ == "__main__":
    main()